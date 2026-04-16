# splitter/utils.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os, zipfile, io

def split_by_sku(file_obj, filename):
    ext = filename.rsplit('.', 1)[-1].lower()
    engine = 'odf' if ext == 'ods' else 'openpyxl'

    df = pd.read_excel(file_obj, engine=engine)
    df.columns = df.columns.str.strip()

    if 'SKU_NO' not in df.columns:
        raise ValueError("Column 'SKU_NO' not found in file.")

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sku, group in df.groupby('SKU_NO'):
            wb = Workbook()
            ws = wb.active
            ws.title = str(sku)[:31]

            # Header styling
            header_fill = PatternFill('solid', start_color='1F4E79')
            header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)

            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row in enumerate(group.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value).font = Font(name='Arial', size=10)

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            safe_name = str(sku).replace('/', '_').replace(' ', '_')
            zf.writestr(f"{safe_name}.xlsx", excel_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer